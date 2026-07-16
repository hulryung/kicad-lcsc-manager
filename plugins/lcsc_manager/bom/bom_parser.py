"""BOM file parsing for batch component import.

Parses a JLCPCB / EasyEDA / KiCad-style BOM file (CSV, or XLSX when the
optional ``openpyxl`` package is available), auto-detects the LCSC part
number column, and returns a de-duplicated list of entries ready for
batch import.

This module is deliberately free of wx / pcbnew dependencies so it can be
unit-tested outside of KiCad.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Sequence

# LCSC part numbers look like ``C`` followed by digits, e.g. ``C2040``,
# ``C25804``. We require at least two digits to avoid matching stray "C1"
# reference designators that leak into a mis-detected column.
LCSC_ID_EXACT_RE = re.compile(r"^C\d{2,}$")

# Header hints (compared against a normalized, alphanumeric-only header).
_DESIGNATOR_EXACT = {"designator", "designators", "reference", "references",
                     "refdes", "ref", "designation"}
_COMMENT_PRIORITY = ("comment", "value", "name", "description", "partname")

# Cells that identify a row as a header row (normalized forms). Used to tell
# a real header apart from title/metadata rows or data rows that merely
# contain the substring "lcsc" (e.g. datasheet URLs).
_HEADER_KEYWORDS = _DESIGNATOR_EXACT | {
    "comment", "value", "name", "description", "partname",
    "footprint", "package", "quantity", "qty", "datasheet",
    "manufacturer", "mfr", "mpn", "manufacturerpart", "supplier",
    "supplierpart",
}


class BomParseError(Exception):
    """Raised when a BOM file cannot be parsed into importable entries."""


@dataclass
class BomEntry:
    """A single de-duplicated component to import."""
    lcsc_id: str
    designators: List[str] = field(default_factory=list)
    quantity: int = 0
    comment: str = ""
    footprint: str = ""


@dataclass
class BomParseResult:
    """Result of parsing a BOM file."""
    entries: List[BomEntry]
    lcsc_column: Optional[str]
    total_data_rows: int
    skipped_rows: int
    skipped_designators: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    @property
    def part_count(self) -> int:
        return len(self.entries)


def _normalize(value: str) -> str:
    """Lower-case and strip everything but ``[a-z0-9]`` for header matching."""
    return re.sub(r"[^a-z0-9]", "", (value or "").lower())


def _split_designators(cell: str) -> List[str]:
    """Split a designator cell (``R1,R2 R3;R4``) into individual refs."""
    if not cell:
        return []
    return [d.strip() for d in re.split(r"[,;\s]+", cell.strip()) if d.strip()]


def _parse_quantity(cell: str) -> Optional[int]:
    """Parse a quantity cell into an int, tolerating ``"3"``/``"3.0"``."""
    if cell is None:
        return None
    cell = str(cell).strip()
    if not cell:
        return None
    try:
        return int(float(cell))
    except (ValueError, TypeError):
        return None


def _decode_bytes(data: bytes) -> str:
    """Decode raw file bytes into text, detecting the encoding.

    Prefers the bundled ``charset_normalizer`` (robust for GBK/Big5 exports
    from Chinese tooling) and falls back to a fixed encoding chain.
    """
    try:
        from charset_normalizer import from_bytes  # bundled in lcsc_manager/lib
        best = from_bytes(data).best()
        if best is not None:
            return str(best).lstrip("﻿")
    except Exception:
        pass

    for enc in ("utf-8-sig", "utf-8", "gbk", "big5", "cp1252", "latin-1"):
        try:
            return data.decode(enc).lstrip("﻿")
        except (UnicodeDecodeError, LookupError):
            continue
    return data.decode("utf-8", errors="replace").lstrip("﻿")


def _sniff_delimiter(sample: str) -> str:
    """Best-effort CSV delimiter detection among comma / semicolon / tab."""
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return dialect.delimiter
    except csv.Error:
        lines = [ln for ln in sample.splitlines() if ln.strip()]
        first = lines[0] if lines else ""
        counts = {d: first.count(d) for d in (",", ";", "\t")}
        best = max(counts, key=counts.get)
        return best if counts[best] else ","


def _read_csv(path: Path) -> List[List[str]]:
    text = _decode_bytes(path.read_bytes())
    if not text.strip():
        raise BomParseError("The BOM file is empty.")
    delimiter = _sniff_delimiter(text[:4096])
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [[("" if c is None else str(c)) for c in row] for row in reader]


def _read_xlsx(path: Path) -> List[List[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise BomParseError(
            "Reading .xlsx BOM files requires the 'openpyxl' package. "
            "Please export your BOM as CSV instead, or install openpyxl."
        )
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c) for c in row])
        return rows
    finally:
        wb.close()


def _looks_like_lcsc_header_cell(cell_norm: str, cell_raw: str) -> bool:
    """True if a cell reads like an 'LCSC Part #'-style column header —
    not a data value (an actual LCSC id) or a URL that contains 'lcsc'."""
    if "lcsc" not in cell_norm:
        return False
    raw = cell_raw.strip()
    if "://" in raw or ".com" in raw.lower():
        return False  # e.g. https://www.lcsc.com/datasheet/C21190.pdf
    if LCSC_ID_EXACT_RE.match(raw.upper()):
        return False  # a data value, not a header
    return len(cell_norm) <= 24  # 'lcscpartnumber' etc., not title sentences


def _find_header_index(rows: Sequence[Sequence[str]]) -> Optional[int]:
    """Locate the header row, or None when the file has no header.

    Some exports prepend metadata/title rows, so a row only counts as the
    header when it convincingly reads like one:
    - an LCSC-header-looking cell plus at least one other known header
      keyword (rules out 'LCSC BOM Export - MyProject' title rows), or
    - two or more known header keywords, or
    - a designator-like plus a footprint-like column.

    Returning None (headerless file, e.g. a bare list of part numbers) makes
    every row a data row instead of silently consuming the first part.
    """
    for idx, row in enumerate(rows[:10]):
        raw = [str(c) for c in row]
        norm = [_normalize(c) for c in raw]

        lcsc_header = any(_looks_like_lcsc_header_cell(n, r)
                          for n, r in zip(norm, raw))
        keyword_hits = sum(1 for n in norm if n in _HEADER_KEYWORDS)
        has_desig = any(c in _DESIGNATOR_EXACT for c in norm)
        has_fp = any(("footprint" in c or "package" in c) for c in norm)

        if lcsc_header and keyword_hits >= 1:
            return idx
        if keyword_hits >= 2:
            return idx
        if has_desig and has_fp:
            return idx
    return None


def _classify_columns(header_norm: List[str]) -> Dict[str, int]:
    """Map roles (designator/quantity/footprint/comment) to column indices."""
    cols: Dict[str, int] = {}

    for idx, h in enumerate(header_norm):
        if "designator" not in cols and h in _DESIGNATOR_EXACT:
            cols["designator"] = idx
        if "quantity" not in cols and ("quantity" in h or h == "qty"):
            cols["quantity"] = idx
        if "footprint" not in cols and ("footprint" in h or "package" in h):
            cols["footprint"] = idx

    # Comment column: pick by priority so "Comment" wins over "Name".
    for key in _COMMENT_PRIORITY:
        for idx, h in enumerate(header_norm):
            if h == key:
                cols["comment"] = idx
                break
        if "comment" in cols:
            break

    return cols


def _find_lcsc_column(header_norm: List[str],
                      data_rows: Sequence[Sequence[str]],
                      classified: Dict[str, int],
                      warnings: List[str]) -> Optional[int]:
    """Find the LCSC part-number column, by header name then by values."""
    # 1) Header-based: any column whose name mentions "lcsc"; prefer the one
    #    that also mentions "part" (e.g. "LCSC Part #").
    lcsc_cols = [i for i, h in enumerate(header_norm) if "lcsc" in h]
    if lcsc_cols:
        for i in lcsc_cols:
            if "part" in header_norm[i]:
                return i
        return lcsc_cols[0]

    # 2) Value-based fallback: among columns NOT already classified as
    #    designator/quantity/footprint/comment, pick the one whose cells most
    #    often look exactly like an LCSC id.
    classified_idx = set(classified.values())
    best_idx, best_hits = None, 0
    for idx in range(len(header_norm)):
        if idx in classified_idx:
            continue
        hits = 0
        for row in data_rows:
            if idx < len(row) and LCSC_ID_EXACT_RE.match(str(row[idx]).strip().upper()):
                hits += 1
        if hits > best_hits:
            best_idx, best_hits = idx, hits

    if best_idx is not None:
        warnings.append(
            "No 'LCSC Part #' header found; auto-detected LCSC part numbers "
            f"in column {best_idx + 1}."
        )
        return best_idx
    return None


def _extract_lcsc_id(cell_value: str) -> str:
    """Extract an LCSC id from a cell only when it appears as a standalone
    token — 'C2040' or 'C2040 (RP2040)' match, but 'STM32C011F4' or
    '1C2040X' must NOT yield a phantom part number."""
    for token in re.split(r"[^A-Za-z0-9]+", cell_value.upper()):
        if LCSC_ID_EXACT_RE.match(token):
            return token
    return ""


def _parse_rows(rows: List[List[str]]) -> BomParseResult:
    rows = [r for r in rows if any((c or "").strip() for c in r)]
    if not rows:
        raise BomParseError("The BOM file contains no data.")

    header_idx = _find_header_index(rows)
    if header_idx is None:
        # Headerless file (e.g. a bare list of part numbers): every row is
        # data. Pad a pseudo-header so column scans cover the widest row.
        header: List[str] = []
        header_norm = [""] * max(len(r) for r in rows)
        data_rows = rows
    else:
        header = rows[header_idx]
        header_norm = [_normalize(c) for c in header]
        data_rows = rows[header_idx + 1:]

    warnings: List[str] = []
    classified = _classify_columns(header_norm)
    lcsc_idx = _find_lcsc_column(header_norm, data_rows, classified, warnings)

    if lcsc_idx is None:
        detected = ", ".join(h for h in header if h.strip()) or "(none)"
        raise BomParseError(
            "Could not find an LCSC part-number column in the BOM.\n"
            f"Detected columns: {detected}\n"
            "Make sure the BOM has an 'LCSC Part #' column."
        )

    desig_idx = classified.get("designator")
    qty_idx = classified.get("quantity")
    comment_idx = classified.get("comment")
    fp_idx = classified.get("footprint")

    def cell(row: Sequence[str], idx: Optional[int]) -> str:
        if idx is None or idx >= len(row):
            return ""
        return str(row[idx]).strip()

    by_id: Dict[str, BomEntry] = {}
    order: List[str] = []
    skipped_rows = 0
    skipped_designators: List[str] = []

    for row in data_rows:
        lcsc_id = _extract_lcsc_id(cell(row, lcsc_idx))
        designators = _split_designators(cell(row, desig_idx))

        if not lcsc_id:
            skipped_rows += 1
            skipped_designators.extend(designators)
            continue

        qty = _parse_quantity(cell(row, qty_idx))
        if qty is None:
            qty = len(designators) if designators else 1

        if lcsc_id in by_id:
            entry = by_id[lcsc_id]
            entry.designators.extend(designators)
            entry.quantity += qty
        else:
            entry = BomEntry(
                lcsc_id=lcsc_id,
                designators=list(designators),
                quantity=qty,
                comment=cell(row, comment_idx),
                footprint=cell(row, fp_idx),
            )
            by_id[lcsc_id] = entry
            order.append(lcsc_id)

    entries = [by_id[i] for i in order]
    return BomParseResult(
        entries=entries,
        lcsc_column=header[lcsc_idx] if lcsc_idx < len(header) else None,
        total_data_rows=len(data_rows),
        skipped_rows=skipped_rows,
        skipped_designators=skipped_designators,
        warnings=warnings,
    )


def parse_bom(file_path) -> BomParseResult:
    """Parse a BOM file into a :class:`BomParseResult`.

    Args:
        file_path: Path to a ``.csv`` (or ``.xlsx``/``.xlsm``) BOM file.

    Returns:
        A :class:`BomParseResult` with de-duplicated importable entries.

    Raises:
        BomParseError: If the file cannot be read or has no LCSC column.
    """
    path = Path(file_path)
    if not path.exists():
        raise BomParseError(f"BOM file not found: {path}")

    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        rows = _read_xlsx(path)
    else:
        rows = _read_csv(path)

    return _parse_rows(rows)
